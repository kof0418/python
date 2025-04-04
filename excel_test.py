from openpyxl import load_workbook

wb = load_workbook("Dodgers.xlsx")
result = []

ws = wb.worksheets[0]
for row in ws.iter_rows():
    result.append([cell.value for cell in row])

# print(result)

sum = 0
for r in result[1:]:
    sum += int(r[11])

print(f"The total number of homeruns for Dodgers was {sum}.")
